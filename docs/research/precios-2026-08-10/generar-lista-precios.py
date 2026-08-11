"""Genera la lista final de precios en Excel para enviar a Felipe."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IVA = 1.19
def neto(p): return p / IVA
def margen(p, c): return (neto(p) - c) / neto(p)

VERDE = '3B5751'; VERDE_SUAVE = 'E3EAE8'; ROJO = 'B0223F'
TINTA = '1C2420'; GRIS = 'F4F2ED'

wb = openpyxl.Workbook()

# ---------- estilos ----------
def encabezado(ws, fila, titulos, ancho=None):
    for i, t in enumerate(titulos, 1):
        c = ws.cell(row=fila, column=i, value=t)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = PatternFill('solid', fgColor=VERDE)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = Border(bottom=Side('thin', color=TINTA))
    ws.row_dimensions[fila].height = 34
    if ancho:
        for i, w in enumerate(ancho, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def titulo(ws, texto, sub=None):
    ws['A1'] = texto
    ws['A1'].font = Font(bold=True, size=15, color=VERDE)
    if sub:
        ws['A2'] = sub
        ws['A2'].font = Font(size=10, italic=True, color='4C5850')

CLP = '"$"#,##0'
PCT = '0.0%'

# =====================================================================
# HOJA 1 — Lista de precios (tienda propia)
# =====================================================================
ws = wb.active
ws.title = 'Lista de precios'
titulo(ws, 'Pedraza Ilustración — Lista de precios',
       'Tienda propia (Shopify). Precios con IVA incluido. Margen calculado sobre precio neto. 11 de agosto de 2026.')

encabezado(ws, 4,
           ['Producto', 'Precio hoy', 'Propuesta\nFelipe', 'Precio\nsugerido', 'Costo\nunitario',
            'Margen\nsugerido', 'Qué se hace', 'Por qué'],
           [22, 12, 12, 12, 12, 11, 20, 62])

# producto, hoy, propuesta, sugerido, costo, decision, motivo
FILAS = [
 ('Lámina A2', 52990, 36990, 36990, 8357, 'Aplicar la baja',
  'Es donde el mercado respalda la baja. Pasa de costar 1,8 veces por cm² lo que cobra la competencia por un formato mayor, a 1,3 veces.'),
 ('Lámina A3', 40990, 23990, 23990, 6261, 'Aplicar la baja',
  'Pasa de +49% a +15% por cm² contra el comparable directo. Se imprime en Chile, así que no depende del dólar ni ocupa capital en un contenedor.'),
 ('Puzzle 1.000 pzs', 29990, 27990, 29990, 7463, 'MANTENER el precio',
  'Bajarlo deja el margen 2,4% sobre julio y cuesta $2,06 millones al año, más de la mitad del costo total de la propuesta. Una tienda ya lo revende a $31.990 con su propio margen encima.'),
 ('Puzzle 60 pzs', None, 18990, 18990, 6599, 'Aplicar (producto nuevo)',
  'Queda en 58,6%, bajo el piso, por el FOB alto de un pedido chico. Se corrige con volumen en el próximo pedido, no con precio.'),
 ('Botella', 32990, 27990, 27990, 9651.16, 'Aplicar la baja',
  'Excepción deliberada al piso de 60%. Categoría muy competitiva: el problema es el costo ($9.651, el más alto del catálogo), no el precio.'),
 ('Naipes', 22990, 22990, 22990, 3770, 'Sin cambio',
  'A −4% de la mediana de mercado. Está donde debe estar.'),
 ('Pin', 7990, 5990, 7990, 1147, 'MANTENER el precio',
  'Era el producto mejor calibrado del catálogo: +7% sobre la referencia de mercado. A $5.990 queda ~20% por debajo y cuesta $333.894 al año sin razón de competencia.'),
 ('Llavero', 9990, 7990, 7990, 1710, 'Aplicar la baja',
  'El margen aguanta de sobra: queda en 74,5%.'),
 ('Totebag', 17990, 15290, 15290, 5080, 'Aplicar la baja',
  'Corregido desde $14.990 para volver sobre el piso de 60%.'),
 ('Libretas', 18990, 16990, 16990, 4200, 'Aplicar la baja',
  'El margen queda en 70,6%. Ojo: el costo de $4.200 es el de la hoja Shopify; la hoja de Mercado Libre dice $7.235 y con ese número el margen sería 54,7%.'),
 ('Postales', 11990, 11990, 11990, 1702, 'Sin cambio', ''),
 ('Calcetines', 9990, 9990, 9990, 1148, 'Sin cambio', ''),
 ('Bandanas', None, 9990, 9990, 1080, 'Aplicar (producto nuevo)', ''),
]

f = 5
for nombre, hoy, prop, sug, costo, decision, motivo in FILAS:
    ws.cell(row=f, column=1, value=nombre).font = Font(bold=True)
    ws.cell(row=f, column=2, value=hoy if hoy else '—').number_format = CLP
    ws.cell(row=f, column=3, value=prop).number_format = CLP
    c = ws.cell(row=f, column=4, value=sug); c.number_format = CLP; c.font = Font(bold=True)
    ws.cell(row=f, column=5, value=costo).number_format = CLP
    m = ws.cell(row=f, column=6, value=margen(sug, costo)); m.number_format = PCT
    if margen(sug, costo) < 0.60:
        m.font = Font(color=ROJO, bold=True)
    d = ws.cell(row=f, column=7, value=decision)
    if decision.startswith('MANTENER'):
        d.font = Font(bold=True, color=ROJO)
        for col in range(1, 9):
            ws.cell(row=f, column=col).fill = PatternFill('solid', fgColor='F8DFE4')
    elif decision == 'Sin cambio':
        d.font = Font(color='4C5850')
    ws.cell(row=f, column=8, value=motivo).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[f].height = 46
    f += 1

f += 1
ws.cell(row=f, column=1, value='Si se aplica todo lo que propone Felipe, la baja cuesta $4.009.412 al año a volumen constante '
        '(sin contar láminas, que no tienen unidades desglosadas). Manteniendo puzzle y pin, cuesta $1.617.255.'
        ).font = Font(italic=True, size=10, color=TINTA)
ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=8)
ws.cell(row=f, column=1).alignment = Alignment(wrap_text=True, vertical='top')
ws.row_dimensions[f].height = 32
ws.freeze_panes = 'A5'

# =====================================================================
# HOJA 2 — Mercado Libre
# =====================================================================
ws2 = wb.create_sheet('Mercado Libre')
titulo(ws2, 'Pedraza Ilustración — Precios de Mercado Libre',
       'Mercado Libre necesita su propio piso de margen: la brecha con la tienda propia va de 12,3 a 23,0 puntos.')
encabezado(ws2, 4,
           ['Producto', 'Tienda\npropia', 'Mercado\nLibre', 'Margen ML\nreportado',
            'Margen ML\ncon envío completo', 'Brecha vs.\ntienda propia', 'Observación'],
           [22, 12, 12, 12, 15, 13, 58])

# nombre, shopify_sugerido, ml, margen_reportado, margen_corregido, margen_shopify, nota
ML = [
 ('Naipes', 22990, 24490, .575, .496, .805, ''),
 ('Puzzle 1.000 pzs', 29990, None, .464, .398, .683, 'Si el precio de la tienda propia queda en $29.990, el de Mercado Libre tiene que subir: a $29.490 quedaría MÁS BARATO que la tienda propia e invertiría la regla de canal. Mínimo $31.490.'),
 ('Botella', 27990, 29490, .394, .328, .590, 'Es el margen más bajo de todo el catálogo en Mercado Libre.'),
 ('Pin', 7990, None, .599, .599, .772, 'Mismo caso que el puzzle: si la tienda propia queda en $7.990, Mercado Libre no puede quedar en $6.490. Mínimo $8.490.'),
 ('Lámina A2', 36990, 39990, .561, .513, .731, ''),
 ('Lámina A3', 23990, 25990, .523, .449, .689, ''),
 ('Llavero', 7990, 8990, .583, .583, .745, 'Bajo $19.990 no hay envío gratis obligatorio, así que el margen reportado no cambia.'),
 ('Libretas', 16990, 18990, .558, .558, .706, 'Bajo $19.990: sin envío que absorber.'),
 ('Totebag', 15290, 16490, .482, .482, .605, 'Bajo $19.990: sin envío que absorber.'),
 ('Puzzle 60 pzs', 18990, 18990, None, None, .586, 'ÚNICO producto con el mismo precio en los dos canales: no recupera nada del costo de Mercado Libre. Con comisión normal el margen cae a 18%-24%. Falta confirmar la comisión del SKU.'),
 ('Postales', 11990, None, None, None, .831, 'No publicado en Mercado Libre.'),
]

f = 5
for nombre, sh, ml, mrep, mcorr, msh, nota in ML:
    ws2.cell(row=f, column=1, value=nombre).font = Font(bold=True)
    ws2.cell(row=f, column=2, value=sh).number_format = CLP
    c = ws2.cell(row=f, column=3, value=ml if ml else 'Recalcular')
    if ml: c.number_format = CLP
    else: c.font = Font(bold=True, color=ROJO)
    if mrep is not None:
        ws2.cell(row=f, column=4, value=mrep).number_format = PCT
        cc = ws2.cell(row=f, column=5, value=mcorr); cc.number_format = PCT
        if mcorr < 0.45: cc.font = Font(color=ROJO, bold=True)
        b = ws2.cell(row=f, column=6, value=msh - mrep); b.number_format = PCT
    else:
        ws2.cell(row=f, column=4, value='Por confirmar').font = Font(color=ROJO)
    ws2.cell(row=f, column=7, value=nota).alignment = Alignment(wrap_text=True, vertical='top')
    ws2.row_dimensions[f].height = 46
    f += 1

f += 1
for linea in [
  'El margen reportado resta solo la mitad del envío de Mercado Libre. El traspaso del 50% ya está dentro del precio,',
  'y a Mercado Libre se le pagan los $3.250 completos: por eso la columna "con envío completo". Pendiente de confirmar.',
  'Los productos bajo $19.990 no cargan envío, porque ahí Mercado Libre no obliga a envío gratis. En esos el número no cambia.']:
    ws2.cell(row=f, column=1, value=linea).font = Font(italic=True, size=10)
    ws2.merge_cells(start_row=f, start_column=1, end_row=f, end_column=7)
    f += 1
ws2.freeze_panes = 'A5'

# =====================================================================
# HOJA 3 — Pendientes
# =====================================================================
ws3 = wb.create_sheet('Pendientes')
titulo(ws3, 'Lo que falta para cerrar la lista', 'Ordenado por cuánto cambia la decisión.')
encabezado(ws3, 4, ['#', 'Qué falta', 'Quién', 'Qué desbloquea'], [5, 46, 14, 60])

PEND = [
 ('El margen de Mercado Libre, ¿resta el envío completo o solo la mitad?', 'Felipe',
  'Cambia entre 4,8 y 7,9 puntos de margen en los 5 productos sobre $19.990. Con envío completo la botella cae a 32,8%.'),
 ('La comisión de Mercado Libre del puzzle de 60 piezas', 'Felipe',
  'Es el único producto sin margen calculable en el canal, y el único que no traspasa nada: hoy vale lo mismo en los dos.'),
 ('El costo de las libretas: $4.200 o $7.235', 'Felipe',
  'Las dos hojas del archivo dan costos distintos para el mismo producto. Con uno margina 70,6% y con el otro 54,7%.'),
 ('Precio de Mercado Libre del puzzle de 1.000 y del pin', 'Los dos',
  'Si la tienda propia mantiene $29.990 y $7.990, los precios de Mercado Libre quedan por debajo e invierten la regla de canal.'),
 ('Las unidades vendidas de láminas', 'Felipe',
  'Es la baja más profunda de la propuesta (−30% y −41,5%) y hoy no se puede costear: no hay unidades desglosadas.'),
 ('¿De dónde sale el 10,1% más de unidades?', 'Los dos',
  'Es lo que exige la baja solo para no perder plata. Hoy no está escrito.'),
 ('El número del piso de margen de Mercado Libre', 'Los dos',
  'Ya se acordó que necesita uno propio. Falta fijarlo, sabiendo que la brecha va de 12,3 a 23,0 puntos.'),
 ('¿Shopify calcula el envío gratis antes o después del descuento?', 'Cualquiera con acceso',
  'Cinco minutos. Define si las promociones se sabotean solas.'),
 ('Arreglar la calculadora de Excel del plan', 'Nosotros',
  'Calcula el margen sobre el precio con IVA: aprueba descuentos de 39,5% cuando el techo real es 28,1%.'),
 ('Las ventas del 1 al 10 de agosto, por producto', 'Felipe',
  'Contesta si la demanda cayó por precio o por el quiebre de stock del puzzle.'),
]
f = 5
for i, (que, quien, desbloquea) in enumerate(PEND, 1):
    ws3.cell(row=f, column=1, value=i).alignment = Alignment(horizontal='center')
    ws3.cell(row=f, column=2, value=que).alignment = Alignment(wrap_text=True, vertical='top')
    ws3.cell(row=f, column=3, value=quien)
    ws3.cell(row=f, column=4, value=desbloquea).alignment = Alignment(wrap_text=True, vertical='top')
    if i <= 3:
        for col in range(1, 5):
            ws3.cell(row=f, column=col).fill = PatternFill('solid', fgColor=VERDE_SUAVE)
    ws3.row_dimensions[f].height = 42
    f += 1
ws3.freeze_panes = 'A5'

RUTA = '/home/user/pedraza-ilustracion/descargas/lista-precios-pedraza.xlsx'
wb.save(RUTA)
print('Escrito', RUTA)

# --- control: releer y verificar los margenes ---
wb2 = openpyxl.load_workbook(RUTA)
w = wb2['Lista de precios']
errores = 0
for row in range(5, 5 + len(FILAS)):
    sug = w.cell(row=row, column=4).value
    costo = w.cell(row=row, column=5).value
    m = w.cell(row=row, column=6).value
    esperado = margen(sug, costo)
    if abs(m - esperado) > 1e-9:
        print('ERROR fila', row); errores += 1
print(f'Verificacion: {len(FILAS)} margenes recalculados, {errores} errores')
print('Hojas:', wb2.sheetnames)
